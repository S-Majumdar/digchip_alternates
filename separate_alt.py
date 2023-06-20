##################################################a
# DIGCHIP.COM Automating Cross Reference for parts
# File name: separate_alt.py
# File type: Python file (.py)
# Author: Sriparno Majumdar
# Date: 18th June 2023
# Description: Automates search for alternate parts on digchip.com
##################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from collections import OrderedDict
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# INPUT FILE
file_path = 'YOUR FILE NAME HERE'
part_num_row = 'YOUR ROW NUMBER HERE'
# PANDAS BLOCK
imp_file = pd.read_excel(file_path)
imp_file[part_num_row].fillna('N-A', inplace = True)
parts = list(imp_file[part_num_row])

# SELENIUM BLOCK
options = EdgeOptions() 
options.add_argument("start-maximized")
#options.add_argument("headless")
driver = webdriver.Edge(options=options)
path = 'https://www.digchip.com/datasheets/cross_reference.php?pn='

# END TIMES
final =[]




# element = driver.find_element('form-control input-sm')
# element = driver.find_element(By.XPATH, '//*[@id="form_cross"]/div/div[1]/input')

# element.submit()




driver.get('https://www.digchip.com/datasheets/cross_reference.php')

# FOR BLOCK

start = len(parts)
print('Finding alternates for %s parts' %(start))
count = 1

for part in parts:

    # driver.get('https://www.digchip.com/datasheets/cross_reference.php')
    
    current = driver.current_url
    print(current)
    if str(current) == 'https://www.digchip.com/datasheets/cross_reference.php' :
        print('At home url')
    else:
        print('At wrong URL, Redirecting')
        driver.get('https://www.digchip.com/datasheets/cross_reference.php')

    time.sleep(4)
    element = driver.find_element(By.XPATH, '//*[@id="form_cross"]/div/div[1]/input').clear()
    time.sleep(3)
    element = driver.find_element(By.XPATH, '//*[@id="form_cross"]/div/div[1]/input')
    element.send_keys(part)

    # Avoiding timeout
    time.sleep(4)

    element.submit()

    # GET CROSS REF SITE
    # driver.get(path + part)

    final_form_dict = {'Part Number': part, 'Alternate MEP_NUMBER': '', 'Alternate Manufacturer': '', 'Cross Type': '', }
    final_form_list = []
    print('Trying part number: ' + part)
    print('Remaining parts: ', (start - count))


    # CRYING BLOCK
    if part == 'N-A':
        print('No part, append None')
        # final.append([part, 'N-A', 'N-A', 'N-A', 'N-A'])
    else:

        try:

            test_element = WebDriverWait(driver, 9).until(EC.presence_of_element_located((By.CLASS_NAME, 'table-responsive')))
            print("Alternates Found")

            table = driver.find_element(By.CLASS_NAME, 'table-responsive')

            rows = table.find_elements(By.TAG_NAME, 'tr')

            # PARA TO DICT
            first_dict = {}
            for row_id, row in enumerate(rows, start=1):
                td_elements = row.find_elements(By.TAG_NAME, 'td')
                if len(td_elements) >= 2:
                    td_element = td_elements[1]
                    paras = td_element.find_elements(By.TAG_NAME, 'p')
                    

                    for para_id, para in enumerate(paras, start=1):
                        key = f'r-{row_id} p-{para_id}' # useless key, for enum test
                        # CHECK FOR LINK
                        try:
                            link = para.find_element(By.TAG_NAME, 'a').get_attribute('href')
                        except NoSuchElementException:
                            link = 'No link found'
                        first_dict[key] = para.text + '\n' + link

            # print(first_dict) # test 1

            # INTERIM DICT
            second_dict = {}
            for key, value in first_dict.items():
                second_dict[key] = value.split('\n')

            # print(second_dict) # tes t 2

            # FINAL FORM DICT/LIST
            for key, values in second_dict.items():
                for value in values:
                    if value.startswith('Cross type:'):
                        cross_type = value.split(': ')[1]
                        part_num = values[0]
                        part_num = part_num.rstrip()
                        mfg = values[1] 
                        link = values[-1]
                        final.append([part, part_num, mfg, cross_type,link])
                        # cmbnd = ', '.join(v for v in values[1:] if not v.startswith('Cross type:'))
                        # final_form_dict[cross_type].append(f'{part_num} - {mfg} \n')

        except TimeoutException:
            print("No Alternates found")
            final.append([part, 'None', 'None', 'None', 'None'])

    # To avoid making too many requests
    count = count + 1

    print("Waiting till next")
    time.sleep(13)
    driver.back()


# FINAL DATAFRAME
print(final) # the end test
final_frame = pd.DataFrame(final, columns = ['MEP_NUMBER', 'Alternate MEP_NUMBER', 'Alternate Manufacturer', 'Cross Type', 'PDF Link'])
final_final_frame = final_frame.sort_values(['MEP_NUMBER', 'Cross Type'], ascending=True)
print(final_final_frame)

# final_frame.to_excel('test_out.xlsx')

# STACKOVERFLOW EXCEL BLOCK

wb = Workbook()
ws = wb.active # to get the actual Worksheet object

# dataframe_to_rows allows to iterate over a dataframe with an interface
# compatible with openpyxl. Each df row will be added to the worksheet.
for r in dataframe_to_rows(final_frame, index=True, header=True):
    ws.append(r)

# iterate over each row and row's cells and apply text wrapping.
for row in ws:
  for cell in row:
    cell.alignment = Alignment(vertical = 'top', wrapText=True)

blod_font = Font(bold= True)
for cell in ws["1:1"]:
    cell.font = blod_font

ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 32
ws.column_dimensions['F'].width = 32



# export the workbook as an excel file.
final_path = 'ALTERNATES_' + file_path
wb.save(final_path)


input("Press enter to exit")
driver.quit()


